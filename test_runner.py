import argparse
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from percy import percy_snapshot

# Define device configurations
DEVICE_CONFIGS = {
    "desktop": None  # Default context (no device emulation)
    # "iphone": "iPhone 13",
    # "ipad": "iPad (gen 7)"
}

def slow_scroll(page, scroll_increment=500, delay=4000):
    """Slowly scrolls down the page to allow lazy-loading elements to appear."""
    scroll_height = page.evaluate("() => document.body.scrollHeight")
    current_position = 0
    while current_position < scroll_height:
        current_position += scroll_increment
        page.evaluate(f"() => window.scrollTo(0, {current_position})")
        # time.sleep(delay)
        page.wait_for_timeout(delay)
        # Recalculate scroll height in case it changes
        scroll_height = page.evaluate("() => document.body.scrollHeight")

def run_test(base_url, path, mode, device_name, accept_button_locator=None):
    # Derive test name from path (remove leading slash and replace slashes with underscores)
    test_name = path.strip("/").replace("/", "_") or "root"

    # Construct the full URL
    url = f"{base_url}{path}"

    with sync_playwright() as p:
        # Get device descriptors
        devices = p.devices

        # Get device configuration
        device_config = devices.get(DEVICE_CONFIGS[device_name]) if DEVICE_CONFIGS[device_name] else None

        # Launch browser
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(**device_config) if device_config else browser.new_context()
        page = context.new_page()

        # Navigate to the URL
        page.goto(url)


        # Set a custom page title for snapshot identification
        device_identifier = device_name.lower().replace(" ", "_")
        page.evaluate(f"() => document.title = 'Snapshot - {test_name} - {device_identifier}'")

        # Click the accept button if it appears
        if accept_button_locator:
            try:
                if page.is_visible(accept_button_locator):
                    page.click(accept_button_locator)
                    time.sleep(1)
            except:
                pass

        # Perform slow scroll
        slow_scroll(page)
        css_selector = "#QSIFeedbackButton-btn"  # Replace this with the actual selector
        page.evaluate(f"""
            const element = document.querySelector('{css_selector}');
            if (element) {{
                element.remove();
            }} else {{
                console.log("Element not found: {css_selector}");
            }}
        """)
        # page.evaluate("document.querySelector('#QSIFeedbackButton-btn').remove();")
        # Generate the snapshot name
        snapshot_name = f"{test_name}_{device_identifier}"
        percy_snapshot(page,snapshot_name)
        print(f"Captured snapshot: {snapshot_name}")

        context.close()
        browser.close()

    return snapshot_name

def read_urls_from_excel(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    paths = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        path = row[0]
        if path:
            paths.append(path)
    return paths

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Visual Testing with Playwright")
    parser.add_argument("--base-url", required=True, help="Base URL of the environment under test")
    parser.add_argument("--threads", type=int, default=2, help="Number of parallel threads")
    parser.add_argument("--mode", choices=["baseline", "validate"], default="validate", help="Test mode: 'baseline' or 'validate'")
    parser.add_argument("--excel-file", default="urls.xlsx", help="Excel file containing paths")
    parser.add_argument("--accept-locator", default=None, help="Locator for the accept button")
    parser.add_argument("--devices", nargs="+", default=["desktop"], help="List of devices to simulate (e.g., desktop, iphone, ipad)")

    args = parser.parse_args()

    # Read paths from Excel
    paths = read_urls_from_excel(args.excel_file)
    if not paths:
        print("No paths found in Excel file.")
        exit(1)

    results = []
    with ThreadPoolExecutor(max_workers=args.threads) as executor:
        futures = {
            executor.submit(
                run_test,
                base_url=args.base_url,
                path=path,
                mode=args.mode,
                device_name=device,
                accept_button_locator=args.accept_locator
            ): (path, device) for path in paths for device in args.devices
        }

        for future in as_completed(futures):
            path, device = futures[future]
            try:
                snapshot_name = future.result()
                print(f"Completed test for {path} on {device}, snapshot: {snapshot_name}")
            except Exception as e:
                print(f"Error in test for {path} on {device}: {e}")
